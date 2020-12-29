""" File to run for whole process """
import time

from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException
from selenium.webdriver.support.ui import WebDriverWait

from settings import Settings


class Automater():
    """ Class which automates Google Clasroom process """

    def __init__(self):
        self.settings = Settings()
        self.set_browser_with_options(self.settings.headless)
        self.wait = WebDriverWait(self.browser, self.settings.max_wait_time)

    def run_process(self):
        """ Run the whole automation process """
        self.login()
        home_link = self.browser.current_url
        classes_length = len(self.wait_for_element(
            'class_links_selector', many=True))
        for i in range(classes_length):
            status = 'teacher'
            class_link = self.wait_for_element(
                'class_links_selector', many=True)[i]
            if class_link.get_attribute('href') in self.settings.student_classes_href:
                status = 'student'
            class_link.click()
            self.go_to_weekly_video_panel()
            data = self.scrape_data(status)
            self.save_data_in_excel(data)
            self.browser.get(home_link)

    def login(self):
        """ Logs in user """
        self.browser.get(f'{self.settings.base_url}h/')
        email_input = self.wait_for_element('email_input')
        email_input.send_keys(self.settings.email_address + Keys.ENTER)
        password_input = self.wait_for_element('password_input')
        password_input.send_keys(self.settings.password + Keys.ENTER)

    def go_to_weekly_video_panel(self):
        """ Takes automation process to the Weekly Video panel """
        self.wait_for_element('work_link').click()
        self.wait_for_element('video_panel_link').click()
        self.scroll_to_bottom()

    def scrape_data(self, status):
        """ Takes the data from the panels """
        videos = self.wait_for_element('videos', many=True)
        data = []
        for video in videos:
            try:
                href = self.wait_for_element(
                    f'{status}_answers_link', video).get_attribute('href')
                video_link = self.wait_for_element(
                    'video_link', video).get_attribute('href')
                res = self.new_window(
                    href, self.get_array_from_video_panel)
                data.append({
                    'answers': res,
                    'href': href,
                    'video_link': video_link
                })
            except (TimeoutException, StaleElementReferenceException):
                pass
        return data

    def save_data_in_excel(self, data):
        """ Writes data to excel """
        try:
            wb = load_workbook('data.xlsx')
        except FileNotFoundError:
            wb = Workbook()
        sheet = wb.active

        for question in data:
            for sub_data in question['answers']:
                sheet.append([*sub_data, question['href'],
                              question['video_link'], *self.settings.defaults])
        sheet.column_dimensions['A'].width = 100
        sheet.column_dimensions['F'].width = 100
        sheet.column_dimensions['G'].width = 100
        sheet.column_dimensions['H'].width = 20
        for x in range(1, sheet.max_row + 1):
            sheet.row_dimensions[x].height = 25
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = cell.alignment.copy(wrapText=True)
        wb.save(filename='data.xlsx')

    # Helper methods
    def set_browser_with_options(self, headless=True):
        """ Returns a headless Chrome """
        options = Options()
        if headless:
            options.add_argument("--headless")
        options.add_argument("--window-size=1920x1080")
        self.browser = webdriver.Chrome(
            executable_path='lib/chromedriver', options=options)

    def wait_for_element(self, selector, item=None, many=False):
        """ Waits for an element if not found, and returns if found """
        selector = self.settings.selector_mapping.get(selector, selector)

        def wrapper(driver):
            if item:
                driver = item
            if many:
                func = driver.find_elements_by_css_selector
            else:
                func = driver.find_element_by_css_selector

            element_or_elements = func(selector)
            if not many:
                element_or_elements = [element_or_elements]
            for element in element_or_elements:
                if not element.is_displayed() or not element.is_enabled():
                    return False
            return element_or_elements if many else element_or_elements[0]

        return self.wait.until(wrapper, f'Can\'t find {selector} in {item if item else "browser"}')

    def tear_down(self):
        """ Quits the browser """
        self.browser.quit()

    def new_window(self, url, func, *args, **kwargs):
        """ Opens a new window at url, calls func, and closes window """
        current_window = self.browser.current_window_handle
        self.browser.execute_script('window.open(arguments[0]);', url)
        new_window = [
            window for window in self.browser.window_handles if window != current_window][0]
        self.browser.switch_to.window(new_window)
        try:
            res = func(*args, **kwargs)
        except (TimeoutException, StaleElementReferenceException) as e:
            self.browser.close()
            self.browser.switch_to.window(current_window)
            raise e

        self.browser.close()
        self.browser.switch_to.window(current_window)
        return res

    def get_array_from_video_panel(self):
        """ Scrapes data from single panel and returns as array """
        answers = self.wait_for_element('answers_selector', many=True)
        res = []
        for answer in answers:
            username = self.wait_for_element('username', answer).text
            date = self.wait_for_element('date', answer).text
            answer = self.wait_for_element('answer', answer).text
            res.append([
                answer,
                date,
                username,
                self.settings.people_data[username][0],
                self.settings.people_data[username][1],
            ])
        return res

    def scroll_to_bottom(self):
        """ Scrolls to the bottom of the page (so that it can fully load) """
        last_height = self.browser.execute_script(
            "return document.body.scrollHeight")

        while True:
            self.browser.execute_script(
                "window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(self.settings.scroll_pause_time)
            new_height = self.browser.execute_script(
                "return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height


if __name__ == '__main__':
    try:
        start_time = time.time()
        automater = Automater()
        automater.run_process()
        print((start_time - time.time()) / 60)
    finally:
        automater.tear_down()
